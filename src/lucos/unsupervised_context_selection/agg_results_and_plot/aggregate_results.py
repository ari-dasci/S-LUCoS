import logging
import re
import numpy as np
import pandas as pd

from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from lucos.utils.paths import results_folder
from lucos.utils.dataset import get_benchmark_by_list_of_datasets
logger = logging.getLogger("lucos")


"""
The results file has these columns:
- dataset
- fold
- instance_selection_space
- instance_selection_algorithm
- evaluator
- n_selected_samples
- metric
- value

The code flow should be:
 - 1: Load the result files (one file per fold) and aggregate them
 - 2: check_columns_and_cast_values
 - 3: load the benchmark and check that results contain all benchmark datasets
 - 4: Filter by selected methods/spaces/evaluators/metrics (ALL_METHOD_SPACE_EVALUATOR_TO_AGGREGATE{SSMA/LUCoS} and METRICS_TO_AGGREGATE)
 - 5: Filter executions that do not satisfy random_state in [BASE_SEED, BASE_SEED + expected_runs - 1]. Warn if any method has fewer executions than expected (NRUNS_PER_SAMPLER)
 - 6: Aggregate all runs of each method/space/evaluator/metric per fold using the mean
 - 7: Aggregate folds of each method/space/evaluator/metric per dataset using the mean. Warn if expected folds are missing (FOLDS_TO_AGGREGATE) for any method/space/evaluator/metric.
"""

MAX_CLASSES = 10
MAX_SAMPLES = None  # Set to an int to aggregate only datasets with n_samples < MAX_SAMPLES.
BASE_SEED = 42
C_MULTIPLIERS = [1, 2, 4, 8, 16, 32]

NRUNS_PER_SAMPLER = {
    "SSMAUnderSampler": 1,
    "RandomUnderSamplerBalanced": 5,
    "RandomUnderSamplerUnsupervised": 5,
    "KMedoidsUnderSamplerHeuristic": 5,
    "KMedoidsUnderSamplerRandom": 5,
    "KMedoidsUnderSamplerK-medoids++Euclidean": 5,
    "KMedoidsUnderSamplerK-medoids++Cosine": 5,
    "RDSSUnderSampler": 1,
    "ZCoreUnderSampler": 1,
}

BASELINE_COLUMN_PREFIX = "OriginalSpace_All_"
EXPERIMENT_ID = 'openml'

RESULTS_LUCoS_EXPERIMENT = f"results_LUCoS_{EXPERIMENT_ID}"
EXPERIMENT_RESULTS_FOLDER = results_folder / RESULTS_LUCoS_EXPERIMENT
RAW_RESULTS_FOLDER = EXPERIMENT_RESULTS_FOLDER / "raw"
AGG_RESULTS_FOLDER = EXPERIMENT_RESULTS_FOLDER / "agg"


def _format_results_path(path: str | Path) -> str:
    """Return paths under results_folder relative to results_folder for compact logs."""
    path = Path(path)
    try:
        return str(path.resolve().relative_to(results_folder.resolve()))
    except ValueError:
        return str(path)


METRICS_TO_AGGREGATE_PER_EVALUATOR =  {"TabPFNv2_5": ["AUC", "ACC", "F1"],
                                       "TabPFNv3": ["AUC", "ACC", "F1"],
                                       "TabICLv2": ["AUC", "ACC", "F1"],
                                       "XGBoost": ["AUC", "ACC", "F1"],
                                       "KNN": ["AUC", "ACC", "F1"],
                                       #"UCD": ["UCD"]
                        }

MAX_MISSING_METRIC_WARNINGS_PER_DATASET_METHOD = 5

METADATA_COLUMNS = [
    "dataset_name",
    "dataset",
    "n_samples",
    "n_features",
    "n_classes",
    "imbalance_ratio",
]

RESULT_KEY_COLUMNS = [
    "dataset",
    "fold",
    "instance_selection_space",
    "instance_selection_algorithm",
    "n_selected_samples",
    "random_state",
    "evaluator",
    "metric",
]

# SSMA
FOLDS_TO_AGGREGATE_SSMA = [0]
ALL_METHODS_SPACES_EVALUATORS_TO_AGGREGATE_SSMA = [
    ("OriginalSpace",    "SSMAUnderSampler",                "TabPFNv2_5"),
    ("OriginalSpace",    "RandomUnderSamplerBalanced",      "TabPFNv2_5"),
    ("OriginalSpace",    "RandomUnderSamplerBalanced",      "KNN"),
    ("OriginalSpace",    "RandomUnderSamplerBalanced",      "XGBoost"),
]

# LUCoS
FOLDS_TO_AGGREGATE_LUCoS = [0,1,2,3,4,5,6,7,8,9]

ALL_METHODS_SPACES_TO_AGGREGATE_LUCoS = [
    ("OriginalSpace",    "RandomUnderSamplerUnsupervised"),
    ("OriginalSpace",    "KMedoidsUnderSamplerK-medoids++Euclidean"),
    ("OriginalSpace",    "KMedoidsUnderSamplerK-medoids++Cosine"),
    ("OriginalSpace",    "RDSSUnderSampler"),
    ("OriginalSpace",    "ZCoreUnderSampler"),
    ("TabClustPFNSpace", "KMedoidsUnderSamplerK-medoids++Euclidean"),
    ("TabClustPFNSpace", "KMedoidsUnderSamplerK-medoids++Cosine"),
    ("TabClustPFNSpace", "RDSSUnderSampler"),
    ("TabClustPFNSpace", "ZCoreUnderSampler"),
]

EVALUATOR_TO_AGGREGATE_LUCoS = ["TabPFNv2_5", "TabPFNv3", "TabICLv2"]#, "UCD"]

ALL_METHODS_SPACES_EVALUATORS_TO_AGGREGATE_LUCoS = [(space, method, evaluator) for space, method in ALL_METHODS_SPACES_TO_AGGREGATE_LUCoS for evaluator in EVALUATOR_TO_AGGREGATE_LUCoS]

def _all_metrics_to_aggregate() -> list[str]:
    """Return all configured metric names preserving declaration order."""
    return list(dict.fromkeys(
        metric
        for metrics in METRICS_TO_AGGREGATE_PER_EVALUATOR.values()
        for metric in metrics
    ))


def _metrics_to_aggregate_for_evaluator(evaluator_name: str) -> list[str]:
    return list(METRICS_TO_AGGREGATE_PER_EVALUATOR.get(str(evaluator_name), []))


def build_output_path(suffix: str) -> Path:
    AGG_RESULTS_FOLDER.mkdir(parents=True, exist_ok=True)
    return AGG_RESULTS_FOLDER / f"{RESULTS_LUCoS_EXPERIMENT}_{suffix}"


def build_aggregated_results_path(section: str) -> Path:
    return build_output_path(f"{section}_agg.csv")


def read_aggregated_results(section: str) -> pd.DataFrame:
    path = build_aggregated_results_path(section)
    if not path.exists():
        raise FileNotFoundError(
            f"Aggregated results not found: {_format_results_path(path)}. "
            "Run unsupervised_context_selection_agg_results.py before this script."
        )
    return pd.read_csv(path)


def _extract_fold_number(path: Path) -> int:
    stem = path.stem
    marker = "_f"
    idx = stem.rfind(marker)
    if idx == -1:
        return 10**9
    fold_str = stem[idx + len(marker):]
    return int(fold_str) if fold_str.isdigit() else 10**9


def _discover_experiment_input_files(experiment_id: str) -> list[Path]:

    pattern = f"{RESULTS_LUCoS_EXPERIMENT}_f*.csv"
    input_paths = [
        p for p in RAW_RESULTS_FOLDER.glob(pattern)
        if "_agg" not in p.stem and _extract_fold_number(p) in set(FOLDS_TO_AGGREGATE_SSMA + FOLDS_TO_AGGREGATE_LUCoS)
    ]
    input_paths = sorted(input_paths, key=_extract_fold_number)

    if input_paths:
        return input_paths

    single_file = RAW_RESULTS_FOLDER / f"{RESULTS_LUCoS_EXPERIMENT}.csv"
    if single_file.exists():
        return [single_file]

    if not input_paths:
        raise FileNotFoundError(
            f"No input files found for experiment '{experiment_id}'. "
            f"Checked pattern '{pattern}' and file '{single_file.name}' in {_format_results_path(RAW_RESULTS_FOLDER)}"
        )
    return input_paths


def _summarize_values(values: pd.Series, max_values: int = 10) -> str:
    unique_values = sorted(set(values.dropna().astype(str)))
    if len(unique_values) <= max_values:
        return ",".join(unique_values)
    return ",".join(unique_values[:max_values]) + f",... (+{len(unique_values) - max_values} more)"


def _warn_if_duplicate_result_keys(df: pd.DataFrame, max_examples: int = 10) -> None:
    missing = [column for column in RESULT_KEY_COLUMNS if column not in df.columns]
    if missing:
        logger.warning("Cannot check duplicate result keys because columns are missing: %s", missing)
        return

    duplicate_mask = df.duplicated(subset=RESULT_KEY_COLUMNS, keep=False)
    duplicate_rows = df.loc[duplicate_mask].copy()
    if duplicate_rows.empty:
        logger.info("No duplicate raw result keys found before aggregation.")
        return

    aggregations = {
        "n_rows": ("value", "size"),
        "n_distinct_values": ("value", lambda values: values.nunique(dropna=False)),
    }
    if "_source_file" in duplicate_rows.columns:
        aggregations["source_files"] = ("_source_file", _summarize_values)

    duplicate_groups = (
        duplicate_rows.groupby(RESULT_KEY_COLUMNS, dropna=False)
        .agg(**aggregations)
        .reset_index()
        .sort_values(["n_distinct_values", "n_rows"], ascending=False)
    )
    conflicting_groups = duplicate_groups[duplicate_groups["n_distinct_values"] > 1]
    identical_groups = duplicate_groups[duplicate_groups["n_distinct_values"] <= 1]

    logger.warning(
        "Found duplicate raw result keys before aggregation: %s duplicate row(s) in %s key group(s). "
        "%s group(s) have conflicting values; %s group(s) have identical values.",
        len(duplicate_rows),
        len(duplicate_groups),
        len(conflicting_groups),
        len(identical_groups),
    )

    algorithm_summary = (
        duplicate_groups.groupby("instance_selection_algorithm", dropna=False)
        .agg(
            duplicate_key_groups=("n_rows", "size"),
            duplicate_rows=("n_rows", "sum"),
            conflicting_value_groups=("n_distinct_values", lambda values: int((values > 1).sum())),
            fold_list=("fold", _summarize_values),
        )
        .reset_index()
        .sort_values(["conflicting_value_groups", "duplicate_key_groups", "duplicate_rows"], ascending=False)
    )
    logger.warning("Duplicate raw result keys by algorithm:\n%s", algorithm_summary.to_string(index=False))
    logger.warning("Example duplicate raw result key groups:\n%s", duplicate_groups.head(max_examples).to_string(index=False))


def _filter_runs_by_expected_seed_and_warn_if_missing(
    df: pd.DataFrame,
    nruns_per_fold: dict[str, int],
) -> pd.DataFrame:
    """Filter runs by expected seed window and warn when a group has fewer runs than expected."""

    out = df.copy()

    # Baseline 'All' and SSMAUnderSampler are evaluated once by design, so it should not be validated
    methods_mask = (out["instance_selection_algorithm"].astype(str) != "All") & \
                   (out["instance_selection_algorithm"].astype(str) != "SSMAUnderSampler")
    method_rows = out.loc[methods_mask].copy()
    if method_rows.empty:
        return out

    # 1) FILTER runs with random_state in [BASE_SEED, BASE_SEED + expected_runs - 1]
    # get expected runs per method
    method_rows["_expected_runs"] = (
        method_rows["instance_selection_algorithm"].astype(str).map(lambda m: int(nruns_per_fold.get(m, 1)))
    )
    method_rows["_max_allowed_seed"] = BASE_SEED + method_rows["_expected_runs"] - 1
    method_rows["_min_allowed_seed"] = BASE_SEED
    # filter by min_seed < random_state < max_seed
    keep_mask = (method_rows["random_state"] >= method_rows["_min_allowed_seed"]) & (method_rows["random_state"] <= method_rows["_max_allowed_seed"])
    filtered_methods = method_rows.loc[keep_mask].drop(columns=["_expected_runs", "_max_allowed_seed"])

    out = pd.concat([out.loc[~methods_mask], filtered_methods], ignore_index=True)

    # 2) WARN if any group has fewer runs than expected after filtering

    group_cols = [
        "dataset",
        "fold",
        "instance_selection_space",
        "instance_selection_algorithm",
        "evaluator",
    ]

    expected_groups = method_rows[group_cols].drop_duplicates().copy()
    expected_groups["expected_runs"] = (
        expected_groups["instance_selection_algorithm"].astype(str).map(lambda m: int(nruns_per_fold.get(m, 1)))
    )

    runs_after_filter = (
        filtered_methods.groupby(group_cols, dropna=False)["random_state"]
        .nunique(dropna=False)
        .reset_index(name="n_runs")
    )
    run_check = expected_groups.merge(runs_after_filter, on=group_cols, how="left")
    run_check["n_runs"] = run_check["n_runs"].fillna(0).astype(int)

    problematic_groups = run_check[run_check["n_runs"] < run_check["expected_runs"]]
    for _, row in problematic_groups.iterrows():
        method_name = str(row["instance_selection_algorithm"])
        evaluator_name = str(row["evaluator"])
        expected_runs = int(row["expected_runs"])
        n_runs = int(row["n_runs"])
        max_allowed_seed = BASE_SEED + expected_runs - 1
        group_desc = ", ".join(f"{col}={row[col]}" for col in group_cols)
        logger.warning(
            "Expected at least %s runs for method '%s' after filtering random_state <= %s, but found %s for %s",
            expected_runs,
            method_name,
            max_allowed_seed,
            n_runs,
            group_desc,
        )

    return out


def _warn_if_benchmark_datasets_are_missing(
    datasets_in_results: list,
    benchmark,
    max_names_in_warning: int = 15,
) -> None:
    """Warn if some datasets from the detected benchmark are not present in the results."""
    datasets_in_bechmark = set(benchmark.datasets_info_filtered["dataset_name"].astype(str).values)
    result_dataset_names = set(pd.Series(datasets_in_results, dtype="string").dropna().astype(str).values)
    missing_dataset_names = sorted(datasets_in_bechmark - result_dataset_names)

    logger.info(f"Datasets in results: {len(result_dataset_names)}, datasets in benchmark: {len(datasets_in_bechmark)}, missing datasets: {len(missing_dataset_names)}")

    if not missing_dataset_names:
        return

    preview = ", ".join(missing_dataset_names[:max_names_in_warning])
    suffix = "" if len(missing_dataset_names) <= max_names_in_warning else ", ..."
    logger.warning(
        "Detected benchmark '%s' has %s valid datasets, but results contain %s. Missing %s dataset(s): %s%s",
        getattr(benchmark, "suitename", "unknown"),
        len(datasets_in_bechmark),
        len(result_dataset_names),
        len(missing_dataset_names),
        preview,
        suffix,
    )


def _filter_results_by_dataset_metadata(df: pd.DataFrame, benchmark) -> pd.DataFrame:
    """Keep only datasets allowed by the active benchmark metadata filters."""

    if df is None or df.empty:
        return pd.DataFrame() if df is None else df.copy()

    out = df.copy()
    out["dataset"] = out["dataset"].astype(str)
    n_datasets_before_filter = out["dataset"].nunique()

    allowed_dataset_names = set(benchmark.datasets_info_filtered["dataset_name"].astype(str))
    keep_mask = out["dataset"].isin(allowed_dataset_names)

    if MAX_SAMPLES is not None and "n_samples" in out.columns:
        raw_n_samples = pd.to_numeric(out["n_samples"], errors="coerce")
        keep_mask = keep_mask & (raw_n_samples < MAX_SAMPLES)

    out = out.loc[keep_mask].copy()
    n_datasets_after_filter = out["dataset"].nunique()

    logger.info(
        "Dataset metadata filter active: kept %s/%s datasets (MAX_CLASSES=%s, MAX_SAMPLES=%s)",
        n_datasets_after_filter,
        n_datasets_before_filter,
        MAX_CLASSES,
        MAX_SAMPLES,
    )

    if n_datasets_after_filter < n_datasets_before_filter:
        kept_dataset_names = set(out["dataset"].astype(str))
        removed_dataset_names = sorted(set(df["dataset"].astype(str)) - kept_dataset_names)
        preview = ", ".join(removed_dataset_names[:15])
        suffix = "" if len(removed_dataset_names) <= 15 else ", ..."
        logger.info(
            "Filtered out %s dataset(s) by metadata: %s%s",
            len(removed_dataset_names),
            preview,
            suffix,
        )

    return out


def _filter_results_by_space_method_evaluator_and_fold(
    df: pd.DataFrame,
    selected_space_method_evaluator_tuples: list[tuple[str, str, str]],
    metrics_to_keep: dict[str, list[str]],
    folds_to_keep: list[int]
) -> pd.DataFrame:
    """Keep baseline rows plus selected method/space/evaluator rows only.
    
    Args:
        df: Input dataframe
        selected_space_method_evaluator_tuples: List of (space, method, evaluator) tuples to keep
        metrics_to_keep: Evaluator -> metric names mapping. If None, keeps all metrics.
    """

    df = df.copy()

    df = df[df["fold"].isin(folds_to_keep)]

    baseline_rows = df[
        (df["instance_selection_space"].astype(str) == "OriginalSpace")
        & (df["instance_selection_algorithm"].astype(str) == "All")
    ]

    selected_df = pd.DataFrame(
        selected_space_method_evaluator_tuples,
        columns=["instance_selection_space", "instance_selection_algorithm", "evaluator"],
    )
    method_rows = df.merge(
        selected_df,
        on=["instance_selection_space", "instance_selection_algorithm", "evaluator"],
        how="inner",
    )

    out = pd.concat([baseline_rows, method_rows], ignore_index=True)
    out = out.drop_duplicates()
    
    # Filter by metrics if specified
    if metrics_to_keep is not None:
        out["evaluator"] = out["evaluator"].astype(str)
        out["metric"] = out["metric"].astype(str)
        evaluator_metrics_df = pd.DataFrame(
            [
                {"evaluator": evaluator_name, "metric": metric_name}
                for evaluator_name, metric_names in metrics_to_keep.items()
                for metric_name in metric_names
            ]
        )
        out = out.merge(
            evaluator_metrics_df,
            on=["evaluator", "metric"],
            how="inner",
        )
    
    return out


def filter_results_by_num_selected_instances(
    df: pd.DataFrame,
    datasets_info: pd.DataFrame,
) -> pd.DataFrame:
    """
    Legacy behavior preserved for backward compatibility.
    Keep only rows whose n_selected_samples matches c * n_classes for each dataset.

    Baseline rows with instance_selection_algorithm == "All" are kept because they are
    needed later to compute OriginalSpace_All_* baseline columns.
    """

    if df is None or df.empty:
        return pd.DataFrame() if df is None else df.copy()

    required_info_cols = {"dataset_name", "n_classes"}
    missing_info_cols = sorted(required_info_cols - set(datasets_info.columns))
    if missing_info_cols:
        raise ValueError(f"Missing required datasets_info columns: {missing_info_cols}")

    out = df.copy()
    out["dataset"] = out["dataset"].astype(str)
    out["n_selected_samples"] = pd.to_numeric(out["n_selected_samples"], errors="coerce")

    dataset_to_n_classes = (
        datasets_info[["dataset_name", "n_classes"]]
        .dropna(subset=["dataset_name", "n_classes"])
        .assign(dataset_name=lambda x: x["dataset_name"].astype(str))
        .drop_duplicates(subset=["dataset_name"])
        .set_index("dataset_name")["n_classes"]
        .astype(int)
        .to_dict()
    )

    baseline_mask = out["instance_selection_algorithm"].astype(str) == "All"
    keep_mask = baseline_mask.copy()

    missing_dataset_names = []
    for dataset_name, dataset_rows in out.loc[~baseline_mask].groupby("dataset", dropna=False):
        dataset_name = str(dataset_name)
        n_classes = dataset_to_n_classes.get(dataset_name)
        if n_classes is None:
            missing_dataset_names.append(dataset_name)
            continue

        allowed_n_selected = {int(c * n_classes) for c in C_MULTIPLIERS}
        dataset_keep_mask = out.index.isin(dataset_rows.index) & out["n_selected_samples"].isin(allowed_n_selected)
        keep_mask = keep_mask | dataset_keep_mask

    if missing_dataset_names:
        preview = ", ".join(sorted(set(missing_dataset_names))[:10])
        suffix = "" if len(set(missing_dataset_names)) <= 10 else ", ..."
        logger.warning(
            "Could not filter n_selected_samples for %s dataset(s) because n_classes was missing: %s%s",
            len(set(missing_dataset_names)),
            preview,
            suffix,
        )

    filtered = out.loc[keep_mask].copy()
    logger.info(
        "Filtered rows by n_selected_samples in C_MULTIPLIERS * n_classes: kept %s/%s rows",
        len(filtered),
        len(out),
    )
    return filtered


def improvability(auc_model, auc_best):
    # The improvability metric I invented
    return (1.0 - (auc_model / auc_best)) * 100


def add_improvability_columns(
    agg_df: pd.DataFrame,
) -> pd.DataFrame:
    """Add {c}C_*_{metric}_I columns using the corresponding OriginalSpace_All_{metric} baseline."""

    if agg_df is None or agg_df.empty:
        return pd.DataFrame() if agg_df is None else agg_df.copy()

    out = agg_df.copy()
    baseline_metrics = [
        str(col)[len(BASELINE_COLUMN_PREFIX):]
        for col in out.columns
        if str(col).startswith(BASELINE_COLUMN_PREFIX)
    ]
    metric_names = list(dict.fromkeys(_all_metrics_to_aggregate() + baseline_metrics))
    metric_names = sorted(metric_names, key=len, reverse=True)

    improvability_columns = {}
    for c in C_MULTIPLIERS:
        prefix = f"{c}C_"
        metric_cols = [
            str(col)
            for col in out.columns
            if str(col).startswith(prefix)
            and not str(col).endswith("_I")
        ]

        for metric_col in metric_cols:
            metric_name = next(
                (candidate for candidate in metric_names if metric_col.endswith(f"_{candidate}")),
                None,
            )
            if metric_name is None:
                continue

            baseline_col = f"{BASELINE_COLUMN_PREFIX}{metric_name}"
            metric_values = pd.to_numeric(out[metric_col], errors="coerce")
            if baseline_col in out.columns:
                baseline_values = pd.to_numeric(out[baseline_col], errors="coerce")
            else:
                baseline_values = pd.Series(np.nan, index=out.index)

            improvability_columns[f"{metric_col}_I"] = improvability(metric_values, baseline_values)

    if improvability_columns:
        out = out.drop(columns=[col for col in improvability_columns if col in out.columns])
        out = pd.concat([out, pd.DataFrame(improvability_columns, index=out.index)], axis=1)

    return out


def add_rank_columns(
    agg_df: pd.DataFrame,
) -> pd.DataFrame:
    """Add rank columns by ranking each metric across methods per dataset and evaluator."""

    if agg_df is None or agg_df.empty:
        return pd.DataFrame() if agg_df is None else agg_df.copy()

    out = agg_df.copy()
    rank_columns = {}

    metric_cols_by_rank_group: dict[tuple[int, str, str], list[str]] = {}
    for column_name in out.columns:
        column_data = get_data_from_column(str(column_name))
        if column_data is None or column_data["suffix"]:
            continue

        group_key = (
            column_data["c_multiplier"],
            column_data["evaluator_name"],
            column_data["metric_name"],
        )
        metric_cols_by_rank_group.setdefault(group_key, []).append(str(column_name))

    for (_, _, _), metric_cols in metric_cols_by_rank_group.items():
        metric_values = out[metric_cols].apply(pd.to_numeric, errors="coerce")
        # All metrics are assumed to be higher-is-better, so descending values get better ranks.
        ranks = metric_values.rank(axis=1, ascending=False, method="dense", na_option="keep")
        for metric_col in metric_cols:
            rank_columns[f"{metric_col}_RANK"] = ranks[metric_col].astype("Int64")

    if rank_columns:
        out = out.drop(columns=[col for col in rank_columns if col in out.columns])
        out = pd.concat([out, pd.DataFrame(rank_columns, index=out.index)], axis=1)

    return out


def _aggregate_metrics_by_runs_and_folds(
    df: pd.DataFrame,
    warn_on_missing: bool = False,
) -> pd.DataFrame:
    """
    Aggregate metrics by runs (group by runs, filter seeds), then by folds.
    Returns merged dataframe with all metrics aggregated.
    
    Args:
        df: Input dataframe
        metrics_in_data: List of metrics to aggregate
        warn_on_missing: If True, warn when a metric has no data. If False, silently skip.
    """
    aggregated_by_metric = {}
    
    for metric_name in _all_metrics_to_aggregate():
        df_metric = df[df["metric"].astype(str) == str(metric_name)].copy()
        allowed_evaluators = [
            evaluator_name
            for evaluator_name, evaluator_metrics in METRICS_TO_AGGREGATE_PER_EVALUATOR.items()
            if metric_name in evaluator_metrics
        ]
        df_metric = df_metric[df_metric["evaluator"].astype(str).isin(allowed_evaluators)]
        df_metric = df_metric.dropna(subset=["dataset", "fold", "instance_selection_space", "instance_selection_algorithm", "n_selected_samples", "value"])
        
        if df_metric.empty:
            if warn_on_missing:
                logger.warning("No data found for metric %s after dropna", metric_name)
            continue
        
        # Rename value to metric_name for internal consistency
        # create one column per metric
        df_metric[metric_name] = df_metric["value"]

        # Aggregate by runs: group and filter expected seed window
        run_group_cols = [
            "dataset",
            "fold",
            "instance_selection_space",
            "instance_selection_algorithm",
            "evaluator",
            "n_selected_samples",
        ]
        if "random_state" in df_metric.columns:
            run_group_cols.insert(2, "random_state")
        for optional_col in ["n_samples", "n_features"]:
            if optional_col in df_metric.columns:
                run_group_cols.append(optional_col)

        df_metric = df_metric.groupby(run_group_cols, dropna=False, as_index=False)[metric_name].mean()
        df_metric = _filter_runs_by_expected_seed_and_warn_if_missing(df_metric, nruns_per_fold=NRUNS_PER_SAMPLER)

        # Aggregate by folds
        fold_group_cols = [
            "dataset",
            "fold",
            "instance_selection_space",
            "instance_selection_algorithm",
            "evaluator",
            "n_selected_samples",
        ]
        for optional_col in ["n_samples", "n_features"]:
            if optional_col in df_metric.columns:
                fold_group_cols.append(optional_col)

        df_metric = df_metric.groupby(fold_group_cols, dropna=False, as_index=False)[metric_name].mean()
        aggregated_by_metric[metric_name] = df_metric
    
    metrics_in_data = list(aggregated_by_metric.keys())

    # Merge all metrics back into a single dataframe
    merged_df = aggregated_by_metric[metrics_in_data[0]].copy() if metrics_in_data else pd.DataFrame()
    for metric_name in metrics_in_data[1:]:
        merge_cols = ["dataset", "fold", "instance_selection_space", "instance_selection_algorithm", "evaluator", "n_selected_samples"]
        for optional_col in ["n_samples", "n_features"]:
            if optional_col in merged_df.columns and optional_col in aggregated_by_metric[metric_name].columns:
                merge_cols.append(optional_col)
        merged_df = merged_df.merge(aggregated_by_metric[metric_name], on=merge_cols, how="outer")
    
    return merged_df

def _ensure_all_metric_combinations(
    df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Detect missing metrics per dataset/method and warn. 
    Returns df with only rows that exist in the data (no materialization of all combinations).
    
    Args:
        df: Aggregated dataframe (output from _aggregate_metrics_by_runs_and_folds). There is a column per metric
        metrics_to_process: List of metric names that should be present
        all_method_space_evaluator_tuples: List of (space, method, evaluator) tuples
    
    Returns:
        Same dataframe df with warnings emitted for missing metrics
    """
    
    method_mask = df["instance_selection_algorithm"].astype(str) != "All"
    if not method_mask.any():
        return df

    missing_by_dataset_space_method_evaluator: dict[tuple[str, str, str, str], dict[str, list[int]]] = {}

    metrics_in_data = [col for col in df.columns if col in _all_metrics_to_aggregate()]

    # Only inspect rows where a metric is actually missing.
    for metric in metrics_in_data:
        if metric not in df.columns:
            continue

        evaluators_for_metric = {
            evaluator_name
            for evaluator_name, evaluator_metrics in METRICS_TO_AGGREGATE_PER_EVALUATOR.items()
            if metric in evaluator_metrics
        }
        evaluator_mask = df["evaluator"].astype(str).isin(evaluators_for_metric)

        missing_rows = df.loc[method_mask & evaluator_mask & df[metric].isna(), 
                              ["dataset", "instance_selection_space", "instance_selection_algorithm", "evaluator", "fold"]]
        if missing_rows.empty:
            continue

        grouped = missing_rows.groupby(
            ["dataset", "instance_selection_space", "instance_selection_algorithm", "evaluator"], 
            dropna=False
        )["fold"].agg(
            lambda folds: sorted({int(f) for f in folds.dropna().astype(int).unique().tolist()})
        )

        for (dataset, space, method, evaluator), folds in grouped.items():
            missing_by_dataset_space_method_evaluator.setdefault(
                (str(dataset), str(space), str(method), str(evaluator)), {}
            )[metric] = folds

    for (dataset, space, method, evaluator), metrics_and_folds in sorted(missing_by_dataset_space_method_evaluator.items()):
        parts = []
        for metric in sorted(metrics_and_folds.keys()):
            folds_list = metrics_and_folds[metric]
            parts.append(f"{metric} in folds {folds_list}")
        details = "; ".join(parts)
        logger.warning(
            "Missing metrics dataset=%s, space=%s, method=%s, evaluator=%s: %s",
            dataset,
            space,
            method,
            evaluator,
            details,
        )
    
    return df

def _aggregate_baselines_for_dataset(
    all_df: pd.DataFrame,
    metrics_to_process: list[str],
    folds_to_aggregate: list[int],
    dataset_name: str,
) -> dict[str, float]:
    """
    Compute baseline metrics (OriginalSpace_All_{metric}) for a dataset.
    Returns dict mapping baseline_col -> aggregated value.
    Warnings only if we expect certain folds but they're missing from data.
    """
    baseline_values = {}
    for metric_name in metrics_to_process:
        baseline_col = f"{BASELINE_COLUMN_PREFIX}{metric_name}"
        if metric_name not in all_df.columns:
            baseline_values[baseline_col] = np.nan
            continue

        all_by_fold = all_df.groupby("fold", dropna=False)[metric_name].mean()
        all_by_fold = all_by_fold[all_by_fold.index.isin(folds_to_aggregate)]
        
        # Only warn if we're missing folds from the expected set
        if len(all_by_fold) < len(folds_to_aggregate):
            valid_folds = sorted(int(fold) for fold in all_by_fold.index.tolist())
            missing_folds = [fold for fold in folds_to_aggregate if fold not in set(valid_folds)]
            logger.warning(
                "%s - OriginalSpace - All - %s: expected folds %s, but only found %s. Using available folds for baseline. Missing folds: %s",
                dataset_name,
                metric_name,
                folds_to_aggregate,
                valid_folds,
                missing_folds,
            )
        baseline_values[baseline_col] = float(all_by_fold.mean()) if not all_by_fold.empty else np.nan
    return baseline_values



def check_columns_and_cast_values(df: pd.DataFrame) -> pd.DataFrame:

    required = {
        "dataset",
        "fold",
        "instance_selection_space",
        "instance_selection_algorithm",
        "n_selected_samples",
        "metric",
        "value",
    }
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    df = df.copy()
    df["dataset"] = df["dataset"].astype(str)
    df["fold"] = pd.to_numeric(df["fold"], errors="coerce")
    df["n_selected_samples"] = pd.to_numeric(df["n_selected_samples"], errors="coerce")
    df["value"] = pd.to_numeric(df["value"], errors="coerce")
    if "random_state" in df.columns:
        df["random_state"] = pd.to_numeric(df["random_state"], errors="coerce")
    df["fold"] = df["fold"].astype(int)

    return df


def aggregate_results(
    raw_df: pd.DataFrame,
    folds_to_aggregate: list[int],
    all_method_space_evaluator_tuples: list[tuple[str, str, str]],
) -> pd.DataFrame:
    """
    Aggregate results for all metrics found in raw_df.
    Creates columns {c}C_{space}_{method}_{metric}. Improvability columns are added by add_improvability_columns.
    """

    if raw_df is None or raw_df.empty:
        return pd.DataFrame()

    df = raw_df.copy()

    df = check_columns_and_cast_values(df)

    # Check results contain all benchmark datasets
    datasets_in_results = sorted(df["dataset"].astype(str).unique())
    logger.info("Aggregate dataset filters: MAX_CLASSES=%s, MAX_SAMPLES=%s", MAX_CLASSES, MAX_SAMPLES)
    if MAX_SAMPLES is not None:
        logger.warning(
            "Aggregating only datasets with n_samples < %s. This is intended for partial results only, "
            "not for the normal full-result aggregation.",
            MAX_SAMPLES,
        )
    benchmark = get_benchmark_by_list_of_datasets(
        datasets_in_results,
        max_classes=MAX_CLASSES,
        max_samples=MAX_SAMPLES,
        only_classification=True,
    )
    _warn_if_benchmark_datasets_are_missing(datasets_in_results, benchmark=benchmark)

    df = _filter_results_by_dataset_metadata(df, benchmark)
    if df.empty:
        logger.warning(
            "No result rows remain after dataset metadata filtering (MAX_CLASSES=%s, MAX_SAMPLES=%s).",
            MAX_CLASSES,
            MAX_SAMPLES,
        )
        return pd.DataFrame()

    df = filter_results_by_num_selected_instances(
        df,
        benchmark.datasets_info_filtered,
    )

    # Step 4: Keep only the selected space/method/evaluator/metric combinations.
    df = _filter_results_by_space_method_evaluator_and_fold(
        df, 
        all_method_space_evaluator_tuples,
        metrics_to_keep=METRICS_TO_AGGREGATE_PER_EVALUATOR,
        folds_to_keep=folds_to_aggregate
    )
    
    # Verify that selected metrics are actually present after filtering
    metrics_in_data = [
        metric
        for metric in _all_metrics_to_aggregate()
        if metric in set(df["metric"].astype(str).unique())
    ]
    
    # Step 5-6: Filter by random_state and aggregate metrics by runs and folds
    df = _aggregate_metrics_by_runs_and_folds(df, warn_on_missing=False)
    
    # Step 7: Ensure all metric combinations are present (fill with NaN where needed)
    df = _ensure_all_metric_combinations(df)

    folds_to_aggregate = [fold for fold in folds_to_aggregate if fold < benchmark.n_folds]
    logger.info("Folds for aggregation: %s", folds_to_aggregate)
    
    # Add dataset metadata
    meta_info = benchmark.datasets_info[["dataset_name", "n_samples", "n_features", "task_type", "n_classes", "imbalance_ratio"]]
    meta_info = meta_info.rename(columns={"dataset_name": "dataset"})
    df = df.merge(meta_info, on="dataset", how="left")

    methods_df = df[df["instance_selection_algorithm"].astype(str) != "All"]
    method_pairs = (
        methods_df[["instance_selection_space", "instance_selection_algorithm", "evaluator"]]
        .drop_duplicates()
        .sort_values(["instance_selection_space", "instance_selection_algorithm", "evaluator"])
        .itertuples(index=False, name=None)
    )
    method_pairs = list(method_pairs)

    # Extract preferred order from all_method_space_evaluator_tuples
    preferred_order = [(space, method, evaluator) for (space, method, evaluator) in all_method_space_evaluator_tuples]

    method_pairs_set = set(method_pairs)
    ordered_method_pairs = [pair for pair in preferred_order if pair in method_pairs_set]
    ordered_method_pairs.extend([pair for pair in method_pairs if pair not in set(ordered_method_pairs)])

    # Build dynamic column names for all selected metrics per method.
    dynamic_columns = []
    for multiplier in C_MULTIPLIERS:
        for space_name, method_name, evaluator_name in ordered_method_pairs:
            for metric_name in _metrics_to_aggregate_for_evaluator(evaluator_name):
                if metric_name not in metrics_in_data:
                    continue
                dynamic_columns.append(f"{multiplier}C_{space_name}_{method_name}_{evaluator_name}_{metric_name}")

    rows = []
    for dataset_name in df["dataset"].astype(str).unique():

        df_dataset = df[df["dataset"].astype(str) == dataset_name]
        n_classes = int(df_dataset["n_classes"].iloc[0])

        row = {
            "dataset": dataset_name,
            "n_samples": int(df_dataset["n_samples"].iloc[0]),
            "n_features": int(df_dataset["n_features"].iloc[0]),
            "n_classes": n_classes,
            "imbalance_ratio": float(df_dataset["imbalance_ratio"].iloc[0]),
        }

        all_mask = (
            (df_dataset["instance_selection_space"].astype(str) == "OriginalSpace")
            & (df_dataset["instance_selection_algorithm"].astype(str) == "All")
        )
        all_df = df_dataset.loc[all_mask].copy()

        # Aggregate baseline metrics for this dataset
        baseline_dict = _aggregate_baselines_for_dataset(
            all_df=all_df,
            metrics_to_process=metrics_in_data,
            folds_to_aggregate=folds_to_aggregate,
            dataset_name=dataset_name,
        )
        row.update(baseline_dict)

        # Cache per-method aggregates once per dataset to avoid repeating the same filter/groupby
        # for each C multiplier. Also track completely missing method/evaluator combinations.
        pair_agg_cache: dict[tuple[str, str, str], pd.DataFrame] = {}
        missing_pairs_for_dataset: list[tuple[str, str, str]] = []

        for space_name, method_name, evaluator_name in ordered_method_pairs:
            df_dataset_space_method = df_dataset[
                (df_dataset["instance_selection_space"].astype(str) == str(space_name))
                & (df_dataset["instance_selection_algorithm"].astype(str) == str(method_name))
                & (df_dataset["evaluator"].astype(str) == str(evaluator_name))
            ]

            if df_dataset_space_method.empty:
                pair_agg_cache[(space_name, method_name, evaluator_name)] = pd.DataFrame()
                missing_pairs_for_dataset.append((space_name, method_name, evaluator_name))
                continue

            pair_metrics = [
                metric_name
                for metric_name in _metrics_to_aggregate_for_evaluator(evaluator_name)
                if metric_name in metrics_in_data
            ]
            if not pair_metrics:
                pair_agg_cache[(space_name, method_name, evaluator_name)] = pd.DataFrame()
                continue

            pair_agg_cache[(space_name, method_name, evaluator_name)] = df_dataset_space_method.groupby(
                ["fold", "n_selected_samples"],
                dropna=False,
            )[pair_metrics].mean().reset_index()

        if missing_pairs_for_dataset:
            preview_parts = [f"{space}/{method}/{evaluator}" for space, method, evaluator in missing_pairs_for_dataset[:8]]
            preview = ", ".join(preview_parts)
            suffix = "" if len(missing_pairs_for_dataset) <= 8 else ", ..."
            logger.warning(
                "No rows found after filtering for dataset=%s in %s method/space/evaluator combinations. They will be NaN. Missing: %s%s",
                dataset_name,
                len(missing_pairs_for_dataset),
                preview,
                suffix,
            )

        for multiplier in C_MULTIPLIERS:
            target_n = int(multiplier * n_classes)

            for space_name, method_name, evaluator_name in ordered_method_pairs:
                pair_agg = pair_agg_cache[(space_name, method_name, evaluator_name)]
                pair_metrics = [
                    metric_name
                    for metric_name in _metrics_to_aggregate_for_evaluator(evaluator_name)
                    if metric_name in metrics_in_data
                ]

                if pair_agg.empty:
                    for metric_name in pair_metrics:
                        metric_col = f"{multiplier}C_{space_name}_{method_name}_{evaluator_name}_{metric_name}"
                        row[metric_col] = np.nan
                    continue

                target_df = pair_agg[pair_agg["n_selected_samples"] == target_n]
                valid_folds = sorted(target_df["fold"].dropna().astype(int).unique().tolist())

                # Skip warning for methods that are evaluated only in specific folds by design.
                if len(valid_folds) < len(folds_to_aggregate) and not (method_name in ["SSMAUnderSampler", "RandomUnderSamplerBalanced"]) and valid_folds == [0]:
                    missing_folds = [fold for fold in folds_to_aggregate if fold not in set(valid_folds)]
                    logger.warning(
                        f"{dataset_name} - {space_name} - {method_name} - C={multiplier} ({target_n}) found values in folds {valid_folds} "
                        f"but expected folds {folds_to_aggregate}. Missing: {missing_folds}"
                    )

                for metric_name in pair_metrics:
                    metric_col = f"{multiplier}C_{space_name}_{method_name}_{evaluator_name}_{metric_name}"

                    row[metric_col] = float(target_df[metric_name].mean()) if not target_df.empty else np.nan

        rows.append(row)

    baseline_cols = [f"{BASELINE_COLUMN_PREFIX}{metric_name}" for metric_name in metrics_in_data]
    out_cols = ["dataset", "n_samples", "n_features", "n_classes", "imbalance_ratio"] + baseline_cols + dynamic_columns
    out_df = pd.DataFrame(rows)
    for col in out_cols:
        if col not in out_df.columns:
            out_df[col] = np.nan
    out_df = out_df[out_cols]

    return out_df


###############################################################################
#                         results table creation                               #
###############################################################################

RESULT_TABLE_NAME_RENAMES = {
    "space": {
        "OriginalSpace": "Original",
        "TabClustPFNSpace": "TabClustPFN",
    },
    "method": {
        "RandomUnderSamplerBalanced": "Random Balanced",
        "RandomUnderSamplerUnsupervised": "Random Unsupervised",
        "KMedoidsUnderSamplerK-medoids++Euclidean": "KMedoids-Euc",
        "KMedoidsUnderSamplerK-medoids++Cosine": "KMedoids-Cos",
        "RDSSUnderSampler": "RDSS",
        "SSMAUnderSampler": "SSMA",
        "ZCoreUnderSampler": "ZCore",
    },
    "evaluator": {
        "TabPFNv2_5": "TabPFNv2_5",
        "TabPFNv3": "TabPFNv3",
        "TabICLv2": "TabICLv2",
        "KNN": "KNN",
        "XGBoost": "XGBoost",
        "UCD": "UCD",
    },
}


def rename_result_table_value(column_type: str, value: str) -> str:
    return RESULT_TABLE_NAME_RENAMES.get(column_type, {}).get(str(value), str(value))


def get_data_from_column(column_name: str) -> dict | None:
    pattern = r"(?P<c_multiplier>\d+)C_(?P<column_body>.+?)(?:_(?P<suffix>I|RANK))?$"
    match = re.fullmatch(pattern, column_name)
    if not match:
        return None

    c_multiplier = int(match.group("c_multiplier"))
    column_body = match.group("column_body")
    suffix = match.group("suffix") or ""

    body_pattern = (
        r"(?P<space_name>[^_]+)_"
        r"(?P<method_name>[^_]+)_"
        r"(?P<evaluator_name>.+)_"
        r"(?P<metric_name>[^_]+)"
    )
    body_match = re.fullmatch(body_pattern, column_body)
    if body_match:
        return {
            "c_multiplier": c_multiplier,
            "space_name": body_match.group("space_name"),
            "method_name": body_match.group("method_name"),
            "evaluator_name": body_match.group("evaluator_name"),
            "metric_name": body_match.group("metric_name"),
            "suffix": suffix,
        }
    return None


def _metric_lower_is_better(metric_name: str) -> bool:
    return str(metric_name).upper().endswith(("_I", "_RANK"))


def _is_c_result_column(column_name: object) -> bool:
    return re.fullmatch(r"\d+C", str(column_name)) is not None


def _latex_escape(value: object) -> str:
    text = "" if pd.isna(value) else str(value)
    replacements = {
        "\\": r"\textbackslash{}",
        "&": r"\&",
        "%": r"\%",
        "$": r"\$",
        "#": r"\#",
        "_": r"\_",
        "{": r"\{",
        "}": r"\}",
        "~": r"\textasciitilde{}",
        "^": r"\textasciicircum{}",
    }
    return "".join(replacements.get(char, char) for char in text)


def _excel_safe_name(value: object, max_length: int = 31) -> str:
    text = str(value)
    text = re.sub(r"[\[\]:*?/\\]", "_", text).strip()
    return (text or "Results")[:max_length]


def _excel_safe_table_name(value: object) -> str:
    text = re.sub(r"\W+", "_", str(value)).strip("_")
    if not text or text[0].isdigit():
        text = f"Table_{text}"
    return text


def format_add_save_result_table_to_excel(results_table: pd.DataFrame, resolved_output_path: Path) -> None:
    """Save one Excel table per metric and color each C column independently."""

    resolved_output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(resolved_output_path, engine="openpyxl") as writer:
        if results_table.empty:
            pd.DataFrame().to_excel(writer, index=False, sheet_name="Results")
            return

        if {"evaluator", "metric"}.issubset(results_table.columns):
            metric_tables = [
                (
                    f"{evaluator_name}_{metric_name}",
                    metric_table.drop(columns=["evaluator", "metric"]).reset_index(drop=True),
                )
                for (evaluator_name, metric_name), metric_table in results_table.groupby(["evaluator", "metric"], sort=False)
            ]
        elif "metric" in results_table.columns:
            metric_tables = [
                (str(metric_name), metric_table.drop(columns=["metric"]).reset_index(drop=True))
                for metric_name, metric_table in results_table.groupby("metric", sort=False)
            ]
        else:
            metric_tables = [("Results", results_table.reset_index(drop=True))]

        if not metric_tables:
            results_table.drop(columns=["metric"], errors="ignore").to_excel(writer, index=False, sheet_name="Results")
            return

        used_sheet_names: set[str] = set()
        used_table_names: set[str] = set()

        for metric_idx, (metric_name, metric_table) in enumerate(metric_tables, start=1):
            base_sheet_name = _excel_safe_name(metric_name)
            sheet_name = base_sheet_name
            suffix = 1
            while sheet_name in used_sheet_names:
                suffix_text = f"_{suffix}"
                sheet_name = f"{base_sheet_name[:31 - len(suffix_text)]}{suffix_text}"
                suffix += 1
            used_sheet_names.add(sheet_name)

            metric_table.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]

            max_row = ws.max_row
            max_col = ws.max_column
            if max_row < 1 or max_col < 1:
                continue

            c_columns = [col for col in metric_table.columns if _is_c_result_column(col)]
            first_c_col_idx = next(
                (idx for idx, col in enumerate(metric_table.columns, start=1) if col in c_columns),
                None,
            )

            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if first_c_col_idx is not None:
                ws.freeze_panes = f"{get_column_letter(first_c_col_idx)}2"

            if max_row >= 2:
                table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
                table_name = _excel_safe_table_name(f"Results_{metric_idx}_{metric_name}")
                while table_name in used_table_names:
                    table_name = f"{table_name}_{len(used_table_names) + 1}"
                used_table_names.add(table_name)
                table = Table(displayName=table_name, ref=table_ref)
                table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleLight9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                ws.add_table(table)

            for col_idx, col_name in enumerate(metric_table.columns, start=1):
                column_letter = get_column_letter(col_idx)
                if col_name not in c_columns:
                    if col_name == "space":
                        ws.column_dimensions[column_letter].width = 20
                    elif col_name == "method":
                        ws.column_dimensions[column_letter].width = 34
                    elif col_name == "evaluator":
                        ws.column_dimensions[column_letter].width = 16
                    else:
                        ws.column_dimensions[column_letter].width = 18
                    continue

                ws.column_dimensions[column_letter].width = 13
                for row_idx in range(2, max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = "0.000"

                if max_row < 3:
                    continue

                valid_excel_rows = [
                    row_idx
                    for row_idx in range(2, max_row + 1)
                    if pd.notna(metric_table.iloc[row_idx - 2, col_idx - 1])
                ]
                if len(valid_excel_rows) < 2:
                    continue

                cell_range = " ".join(f"{column_letter}{row_idx}" for row_idx in valid_excel_rows)
                lower_is_better = _metric_lower_is_better(str(metric_name))
                start_color = "00B050" if lower_is_better else "FF0000"
                end_color = "FF0000" if lower_is_better else "00B050"
                color_scale = ColorScaleRule(
                    start_type="min",
                    start_color=start_color,
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFFF00",
                    end_type="max",
                    end_color=end_color,
                )
                ws.conditional_formatting.add(cell_range, color_scale)


def format_add_save_result_table_to_tex(
    results_table: pd.DataFrame,
    resolved_output_path: Path,
    lower_is_better: bool = False,
) -> None:
    """Save results table to LaTeX and mark best/second-best values per C and metric."""

    resolved_output_path.parent.mkdir(parents=True, exist_ok=True)

    if results_table.empty:
        resolved_output_path.write_text("", encoding="utf-8")
        return

    results_table = results_table.copy()
    random_undersampler_methods = {
        "RandomUnderSamplerBalanced",
        "RandomUnderSamplerUnsupervised",
        "Random Balanced",
        "Random Unsupervised",
    }
    kmedoids_methods = {
        "KMedoids",
        "KMedoidsUnderSamplerRandom",
        "KMedoidsUnderSamplerHeuristic",
        "KMedoidsUnderSamplerK-medoids++Cosine",
        "KMedoidsUnderSamplerK-medoids++Euclidean",
    }
    if {"method", "space"}.issubset(results_table.columns):
        results_table.loc[results_table["method"].isin(random_undersampler_methods), "space"] = "-"
        space_order = {
            space_name: order
            for order, space_name in enumerate(dict.fromkeys(results_table["space"].astype(str).tolist()))
        }
        results_table = (
            results_table.assign(
                _latex_random_order=~results_table["method"].isin(random_undersampler_methods),
                _latex_space_order=results_table["space"].astype(str).map(space_order),
                _latex_kmedoids_order=results_table["method"].isin(kmedoids_methods),
            )
            .sort_values(
                by=["_latex_random_order", "_latex_space_order", "_latex_kmedoids_order"],
                kind="stable",
            )
            .drop(columns=["_latex_random_order", "_latex_space_order", "_latex_kmedoids_order"])
        )

    c_columns = [col for col in results_table.columns if _is_c_result_column(col)]
    formatted_table = results_table.copy()

    for col_name in c_columns:
        formatted_table[col_name] = formatted_table[col_name].map(
            lambda value: "" if pd.isna(value) else f"{float(value):.2f}"
        )

    if "metric" in results_table.columns:
        for _, metric_rows in results_table.groupby("metric", sort=False).groups.items():
            for col_name in c_columns:
                numeric_values = pd.to_numeric(results_table.loc[metric_rows, col_name], errors="coerce").dropna()
                unique_values = sorted(numeric_values.unique(), reverse=not lower_is_better)
                if not unique_values:
                    continue

                best_value = unique_values[0]
                second_value = unique_values[1] if len(unique_values) > 1 else None

                for row_idx in metric_rows:
                    value = pd.to_numeric(results_table.at[row_idx, col_name], errors="coerce")
                    if pd.isna(value):
                        continue

                    formatted_value = f"{float(value):.2f}"
                    if np.isclose(value, best_value, equal_nan=False):
                        formatted_table.at[row_idx, col_name] = rf"\textbf{{{formatted_value}}}"
                    elif second_value is not None and np.isclose(value, second_value, equal_nan=False):
                        formatted_table.at[row_idx, col_name] = rf"\underline{{{formatted_value}}}"

    caption_metric = "Metric"
    is_rank_table = False
    if "metric" in results_table.columns:
        metric_values = results_table["metric"].dropna().astype(str).unique()
        if len(metric_values) == 1:
            metric_value = metric_values[0]
            is_rank_table = metric_value.endswith("_RANK")
            caption_metric = re.sub(r"_RANK$", "", metric_value)

    formatted_table = formatted_table.reset_index(drop=True)
    if "metric" in formatted_table.columns:
        formatted_table = formatted_table.drop(columns=["metric"])
    headers = [_latex_escape(col_name) for col_name in formatted_table.columns]
    vertical_separator_after = {"space", "method", "metric"}
    group_columns = [col for col in ["space", "method"] if col in formatted_table.columns]
    alignment_parts = []
    for col_name in formatted_table.columns:
        if col_name in c_columns:
            alignment_parts.append("r")
        elif col_name in group_columns:
            alignment_parts.append("c")
        else:
            alignment_parts.append("l")

        if col_name in vertical_separator_after:
            alignment_parts.append("|")
    column_alignment = "".join(alignment_parts)
    caption_text = (
        f"{caption_metric} rank ({'lower' if lower_is_better else 'higher'} is better)"
        if is_rank_table
        else f"Mean {caption_metric} ({'lower' if lower_is_better else 'higher'} is better)"
    )
    label_suffix = "rank" if is_rank_table else "mean"
    table_label = f"tab:table_{caption_metric.lower()}_{label_suffix}"
    lines = [
        r"\begin{table}[h]",
        r"\centering",
        rf"\caption{{{_latex_escape(caption_text)}}}",
        rf"\label{{{table_label}}}",
        rf"\begin{{tabular}}{{{column_alignment}}}",
        r"\toprule",
        " & ".join(headers) + r" \\",
        r"\midrule",
    ]

    def _get_contiguous_group_sizes(group_by_columns: list[str]) -> dict[int, int]:
        if not group_by_columns:
            return {}

        group_sizes: dict[int, int] = {}
        current_group = None
        current_start = 0
        for row_idx, row in formatted_table.iterrows():
            row_group = tuple(row[col] for col in group_by_columns)
            if current_group is None:
                current_group = row_group
                current_start = row_idx
                continue
            if row_group != current_group:
                group_sizes[current_start] = row_idx - current_start
                current_group = row_group
                current_start = row_idx
        group_sizes[current_start] = len(formatted_table) - current_start
        return group_sizes

    multirow_sizes_by_column: dict[str, dict[int, int]] = {}
    if "space" in formatted_table.columns:
        multirow_sizes_by_column["space"] = _get_contiguous_group_sizes(["space"])
    if "method" in formatted_table.columns:
        method_group_columns = [col for col in ["space", "method"] if col in formatted_table.columns]
        multirow_sizes_by_column["method"] = _get_contiguous_group_sizes(method_group_columns)

    space_starts = set(multirow_sizes_by_column.get("space", {}).keys())
    n_columns = len(formatted_table.columns)

    for row_idx, row in formatted_table.iterrows():
        if row_idx != 0 and row_idx in space_starts:
            lines.append(rf"\cmidrule(lr){{1-{n_columns}}}")

        row_values = []
        for col_name in formatted_table.columns:
            value = row[col_name]
            if col_name in c_columns:
                row_values.append(str(value))
            elif col_name in multirow_sizes_by_column:
                column_group_sizes = multirow_sizes_by_column[col_name]
                if row_idx in column_group_sizes:
                    row_values.append(rf"\multirow[c]{{{column_group_sizes[row_idx]}}}{{*}}{{{_latex_escape(value)}}}")
                else:
                    row_values.append("")
            else:
                row_values.append(_latex_escape(value))
        lines.append(" & ".join(row_values) + r" \\")

    lines.extend(
        [
            r"\bottomrule",
            r"\end{tabular}",
            r"\end{table}",
            "",
        ]
    )
    resolved_output_path.write_text("\n".join(lines), encoding="utf-8")


def create_results_table(
    agg_df: pd.DataFrame,
    output_path: Path,
    *,
    drop_evaluator_column: bool = False,
    include_evaluator_in_tex_filename: bool = True,
) -> pd.DataFrame:
    """
    Average dataset rows and create a compact results table.

    Rows are space/method/metric combinations. Columns are C-specific values.
    Rank columns such as AUC_RANK are treated as metrics. Improvability columns
    are omitted from this compact table.
    """

    agg_df = agg_df.drop(columns=[col for col in agg_df.columns if str(col).endswith("_I")], errors="ignore")

    records = []
    for column_name in agg_df.columns:
        column_data = get_data_from_column(str(column_name))
        if column_data is None:
            continue

        suffix = column_data["suffix"]
        metric_label = column_data["metric_name"] if not suffix else f"{column_data['metric_name']}_{suffix}"
        records.append(
            {
                "space": rename_result_table_value("space", column_data["space_name"]),
                "method": rename_result_table_value("method", column_data["method_name"]),
                "evaluator": rename_result_table_value("evaluator", column_data["evaluator_name"]),
                "metric": metric_label,
                "result_column": f"{column_data['c_multiplier']}C",
                "value": pd.to_numeric(agg_df[column_name], errors="coerce").mean(),
            }
        )

    if not records:
        results_table = pd.DataFrame()
    else:
        results_table = pd.DataFrame.from_records(records)
        known_column_order = [f"{c_multiplier}C" for c_multiplier in C_MULTIPLIERS]
        present_columns = set(results_table["result_column"])
        column_order = [column_name for column_name in known_column_order if column_name in present_columns]
        column_order.extend(
            column_name
            for column_name in dict.fromkeys(results_table["result_column"].tolist())
            if column_name not in column_order
        )
        results_table = results_table.pivot_table(
            index=["space", "method", "evaluator", "metric"],
            columns="result_column",
            values="value",
            aggfunc="mean",
        )
        results_table = results_table.reindex(columns=column_order).reset_index()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path_xlsx = output_path.with_suffix(".xlsx")
    output_path_tex = output_path.parent / "latex" / output_path.with_suffix(".tex").name

    excel_results_table = (
        results_table.drop(columns=["evaluator"], errors="ignore")
        if drop_evaluator_column
        else results_table
    )
    format_add_save_result_table_to_excel(excel_results_table, output_path_xlsx)

    try:
        # Save each table separately in .tex
        for (evaluator_name, metric_name), results_metric in results_table.groupby(["evaluator", "metric"], sort=False):
            tex_file_parts = [output_path_tex.stem]
            if include_evaluator_in_tex_filename:
                evaluator_file_name = str(evaluator_name)
                for raw_evaluator_name in METRICS_TO_AGGREGATE_PER_EVALUATOR:
                    if rename_result_table_value("evaluator", raw_evaluator_name) == str(evaluator_name):
                        evaluator_file_name = str(raw_evaluator_name)
                        break
                tex_file_parts.append(evaluator_file_name)
            tex_file_parts.append(str(metric_name))
            output_path_tex_metric = output_path_tex.with_name(
                f"{'_'.join(tex_file_parts)}.tex"
            )
            tex_results_metric = (
                results_metric.drop(columns=["evaluator"], errors="ignore")
                if drop_evaluator_column
                else results_metric
            )
            format_add_save_result_table_to_tex(
                tex_results_metric,
                output_path_tex_metric,
                lower_is_better=str(metric_name).endswith("_RANK"),
            )
    except KeyError as e:
        logger.warning(
            "Could not generate metric-specific TeX tables: column '%s' not found in results table. "
            "This may occur if some experiments are missing (e.g., SSMA results). Continuing with main results table.",
            str(e).strip("'"),
        )

    logger.info("Saved results table to %s", _format_results_path(output_path_xlsx))
    return excel_results_table


def _result_table_evaluators(agg_df: pd.DataFrame) -> list[str]:
    """Return evaluator names present in result columns, preserving column order."""

    evaluators = []
    for column_name in agg_df.columns:
        column_data = get_data_from_column(str(column_name))
        if column_data is None:
            continue
        evaluators.append(column_data["evaluator_name"])
    return list(dict.fromkeys(evaluators))


def _filter_result_columns_by_evaluator(agg_df: pd.DataFrame, evaluator_name: str) -> pd.DataFrame:
    """Keep metadata and C-result columns for a single evaluator."""

    columns_to_keep = []
    for column_name in agg_df.columns:
        column_data = get_data_from_column(str(column_name))
        if column_data is None or column_data["evaluator_name"] == evaluator_name:
            columns_to_keep.append(column_name)
    return agg_df.loc[:, columns_to_keep].copy()


def create_results_tables_by_evaluator(
    agg_df: pd.DataFrame,
    output_folder: Path,
    *,
    file_prefix: str = "LUCoS",
) -> dict[str, pd.DataFrame]:
    """Create one compact Excel/LaTeX result table folder per evaluator."""

    output_tables = {}
    for evaluator_name in _result_table_evaluators(agg_df):
        evaluator_folder = output_folder / str(evaluator_name)
        evaluator_agg_df = _filter_result_columns_by_evaluator(agg_df, evaluator_name)
        evaluator_output_path = evaluator_folder / f"{file_prefix}_{evaluator_name}_results_table.xlsx"
        output_tables[evaluator_name] = create_results_table(
            evaluator_agg_df,
            evaluator_output_path,
            drop_evaluator_column=True,
            include_evaluator_in_tex_filename=False,
        )
    return output_tables


def check_missing_values(df: pd.DataFrame) -> None:
    """Print missing cells with row dataset and decoded column metadata when possible."""

    missing_mask = df.isna()
    total_missing = int(missing_mask.to_numpy().sum())
    if total_missing == 0:
        return

    print(f"Found {total_missing} missing cells across {int(missing_mask.any(axis=1).sum())} row(s).")

    for column_name in missing_mask.columns:
        missing_rows = missing_mask.index[missing_mask[column_name]].tolist()
        if not missing_rows:
            continue
        datasets = [str(value) for value in df.loc[missing_rows, "dataset"].dropna().astype(str).unique().tolist()]
        datasets_text = ", ".join(datasets) if datasets else "<unknown>"
        print(f"{column_name.ljust(60)} | datasets missing: {datasets_text}")


###############################################################################
#                     formatted aggregate excel creation                       #
###############################################################################

def format_and_save_agg_excel(out_df: pd.DataFrame, output_path: Path) -> None:
    """Save results to Excel with table formatting, grouped C headers, and colored header cells."""

    output_path.parent.mkdir(parents=True, exist_ok=True)

    metrics_found = []
    for metric_name in _all_metrics_to_aggregate():
        if any(str(col).endswith(f"_{metric_name}") for col in out_df.columns):
            metrics_found.append(metric_name)

    logger.info("Detected metrics in Excel formatter: %s", metrics_found)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Reserve row 1 for merged C-group labels, and place table header at row 2.
        out_df.to_excel(writer, index=False, sheet_name="Results", startrow=1)
        ws = writer.sheets["Results"]

        max_row = ws.max_row
        max_col = ws.max_column
        if max_row < 2 or max_col < 1:
            return

        data_start_row = 3
        data_end_row = max_row
        total_row_idx = data_end_row + 1

        # Add empty total row (table metadata will define formulas/labels).
        for col_idx in range(1, max_col + 1):
            ws.cell(row=total_row_idx, column=col_idx).value = None

        max_row = total_row_idx

        # Create Excel table over the exported dataframe range (header row + data rows + total row).
        table_ref = f"A2:{get_column_letter(max_col)}{max_row}"
        table = Table(displayName="AggregatedResults", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight9",
            showFirstColumn=False,
            showLastColumn=True,
            showRowStripes=True,
            showColumnStripes=False,
        )
        # Configure total row with proper settings
        table.totalsRowShown = True
        table.totalsRowCount = 1
        ws.add_table(table)
        
        # Initialize table columns if not already done
        if not table.tableColumns:
            table._initialise_columns()
        
        # Configure total row metadata for each table column.
        totals_label_set = False
        for col_idx, col_name in enumerate(out_df.columns, start=1):
            if col_idx - 1 < len(table.tableColumns):
                table_col = table.tableColumns[col_idx - 1]
                if (not totals_label_set) and col_name in {"dataset_name", "dataset"}:
                    table_col.totalsRowLabel = "TOTAL_AVG"
                    totals_label_set = True
                elif col_name not in METADATA_COLUMNS:
                    table_col.totalsRowFunction = "average"

        # Add thick borders around total row
        thick_side = Side(style="thick")
        for col_idx in range(1, max_col + 1):
            total_cell = ws.cell(row=total_row_idx, column=col_idx)
            is_first = (col_idx == 1)
            is_last = (col_idx == max_col)
            total_cell.border = Border(
                left=thick_side if is_first else total_cell.border.left,
                right=thick_side if is_last else total_cell.border.right,
                top=thick_side,
                bottom=thick_side,
            )

        # Set width=12 for all non-metadata columns.
        for col_idx, col_name in enumerate(out_df.columns, start=1):
            if col_name not in METADATA_COLUMNS:
                ws.column_dimensions[get_column_letter(col_idx)].width = 12

        # Add merged C-group headers in row 1, spanning contiguous column blocks for each C.
        c_group_bounds: list[tuple[int, int]] = []
        for c in C_MULTIPLIERS:
            col_indices = []
            prefix = f"{c}C_"
            for col_idx, col_name in enumerate(out_df.columns, start=1):
                if col_name.startswith(prefix):
                    col_indices.append(col_idx)

            if not col_indices:
                continue

            col_indices = sorted(col_indices)
            block_start = col_indices[0]
            previous_col = col_indices[0]
            contiguous_blocks: list[tuple[int, int]] = []
            for col_idx in col_indices[1:]:
                if col_idx == previous_col + 1:
                    previous_col = col_idx
                    continue
                contiguous_blocks.append((block_start, previous_col))
                block_start = col_idx
                previous_col = col_idx
            contiguous_blocks.append((block_start, previous_col))

            for start_col, end_col in contiguous_blocks:
                c_group_bounds.append((start_col, end_col))
                if start_col != end_col:
                    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                top_cell = ws.cell(row=1, column=start_col)
                top_cell.value = f"{c}C : {c} * n_classes"
                top_cell.font = Font(size=48, bold=True)
                top_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add thick vertical borders between C groups.
        thick_side = Side(style="thick")
        for start_col, end_col in c_group_bounds:
            for row_idx in range(1, max_row + 1):
                left_cell = ws.cell(row=row_idx, column=start_col)
                right_cell = ws.cell(row=row_idx, column=end_col)

                left_cell.border = Border(
                    left=thick_side,
                    right=left_cell.border.right,
                    top=left_cell.border.top,
                    bottom=left_cell.border.bottom,
                    diagonal=left_cell.border.diagonal,
                    diagonal_direction=left_cell.border.diagonal_direction,
                    vertical=left_cell.border.vertical,
                    horizontal=left_cell.border.horizontal,
                )
                right_cell.border = Border(
                    left=right_cell.border.left,
                    right=thick_side,
                    top=right_cell.border.top,
                    bottom=right_cell.border.bottom,
                    diagonal=right_cell.border.diagonal,
                    diagonal_direction=right_cell.border.diagonal_direction,
                    vertical=right_cell.border.vertical,
                    horizontal=right_cell.border.horizontal,
                )

        # Color table headers (row 2) by method/space conventions.
        fills = {
            "gray": PatternFill(fill_type="solid", fgColor="BFBFBF"),
            "green": PatternFill(fill_type="solid", fgColor="92D050"),
            "orange": PatternFill(fill_type="solid", fgColor="F4B183"),
            "blue": PatternFill(fill_type="solid", fgColor="9DC3E6"),
            "gold": PatternFill(fill_type="solid", fgColor="FFD700"),
        }

        def format_header_name(col_name: str) -> str:
            """Format header name by replacing underscores with newlines for metric columns."""
            col_name_str = str(col_name)
            # Check if it's a metric column starting with digit+C (e.g., "1C_OriginalSpace_...")
            if len(col_name_str) > 2 and col_name_str[0].isdigit() and "C_" in col_name_str[:4]:
                col_name_str = col_name_str.replace("Space", "").replace("UnderSampler", "").replace("Random", "Random\n")
                parts = col_name_str.split("_")
                return "\n".join(parts)
            # For metadata columns keep as-is
            return col_name_str

        for col_idx, col_name in enumerate(out_df.columns, start=1):
            header_cell = ws.cell(row=2, column=col_idx)
            col_name_l = str(col_name).lower()

            # Set formatted header text
            header_cell.value = format_header_name(col_name)
            if col_idx - 1 < len(table.tableColumns):
                table.tableColumns[col_idx - 1].name = str(header_cell.value)

            if "originalspace" in col_name_l and "ssma" in col_name_l:
                header_cell.fill = fills["gold"]
            elif "ssma" in col_name_l or "random" in col_name_l:
                header_cell.fill = fills["gray"]
            elif "tabclustpfn" in col_name_l:
                header_cell.fill = fills["blue"]
            elif "originalspace" in col_name_l and not str(col_name).startswith(BASELINE_COLUMN_PREFIX):
                header_cell.fill = fills["orange"]

            header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Apply numeric format to non-metadata table columns (including TOTAL_AVG formula row).
        for col_idx, col_name in enumerate(out_df.columns, start=1):
            if col_name in METADATA_COLUMNS:
                continue
            number_format = "0" if str(col_name).endswith("_RANK") else "0.000"
            for row_idx in range(3, max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = number_format

        # Apply conditional formatting to imbalance_ratio: green (1) -> yellow (8) -> red (15).
        if "imbalance_ratio" in out_df.columns:
            imbalance_col_idx = out_df.columns.get_loc("imbalance_ratio") + 1
            imbalance_letter = get_column_letter(imbalance_col_idx)
            imbalance_range = f"{imbalance_letter}{data_start_row}:{imbalance_letter}{data_end_row + 1}"
            imbalance_scale = ColorScaleRule(
                start_type="num",
                start_value=1,
                start_color="00B050",
                mid_type="num",
                mid_value=8,
                mid_color="FFFF00",
                end_type="num",
                end_value=15,
                end_color="FF0000",
            )
            ws.conditional_formatting.add(imbalance_range, imbalance_scale)

        # Apply conditional formatting to OriginalSpace_All_* baseline columns: red (0.5) -> yellow (0.75) -> green (1.0).
        baseline_cols = [col for col in out_df.columns if str(col).startswith(BASELINE_COLUMN_PREFIX)]
        for baseline_col_name in baseline_cols:
            baseline_col_idx = out_df.columns.get_loc(baseline_col_name) + 1
            baseline_letter = get_column_letter(baseline_col_idx)
            baseline_range = f"{baseline_letter}{data_start_row}:{baseline_letter}{data_end_row + 1}"
            baseline_scale = ColorScaleRule(
                start_type="num",
                start_value=0.5,
                start_color="FF0000",
                mid_type="num",
                mid_value=0.75,
                mid_color="FFFF00",
                end_type="num",
                end_value=1.0,
                end_color="00B050",
            )
            ws.conditional_formatting.add(baseline_range, baseline_scale)

        ws.row_dimensions[1].height = 60

    logger.info("Saved formatted aggregated results to %s", _format_results_path(output_path))


def read_all_experiment_results(experiment_id: str) -> pd.DataFrame:
    input_paths = _discover_experiment_input_files(experiment_id)
    logger.info("Found %s input files for experiment '%s'", len(input_paths), experiment_id)

    raw_frames = []
    for input_path in input_paths:
        logger.info("Reading results from %s", _format_results_path(input_path))
        df = pd.read_csv(input_path)
        df["_source_file"] = input_path.name
        raw_frames.append(df)

    raw_df = pd.concat(raw_frames, ignore_index=True) if raw_frames else pd.DataFrame()
    return raw_df



def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    )

    raw_df = read_all_experiment_results(EXPERIMENT_ID)
    raw_df = check_columns_and_cast_values(raw_df)
    _warn_if_duplicate_result_keys(raw_df)


    ###############################################################
    #                      SSMA agregation                        #
    ###############################################################

    output_path_raw_ssma = build_aggregated_results_path("SSMA")
    output_path_formated_ssma = build_output_path(f"SSMA_agg_formated.xlsx")

    agg_df_ssma = aggregate_results(raw_df, FOLDS_TO_AGGREGATE_SSMA, ALL_METHODS_SPACES_EVALUATORS_TO_AGGREGATE_SSMA)
    agg_df_ssma = add_improvability_columns(agg_df_ssma)
    agg_df_ssma = add_rank_columns(agg_df_ssma)
    agg_df_ssma.to_csv(output_path_raw_ssma, index=False)
    check_missing_values(agg_df_ssma)
    format_and_save_agg_excel(agg_df_ssma, output_path_formated_ssma)

    logger.info("Saved SSMA aggregated results to %s", _format_results_path(output_path_raw_ssma))
    logger.info("Saved SSMA formatted aggregated results to %s", _format_results_path(output_path_formated_ssma))


    ###############################################################
    #                      LUCoS agregation                        #
    ###############################################################


    output_path_raw_lucos = build_aggregated_results_path("LUCoS")
    output_path_formated_lucos = build_output_path(f"LUCoS_agg_formated.xlsx")

    agg_df_lucos = aggregate_results(raw_df, FOLDS_TO_AGGREGATE_LUCoS, ALL_METHODS_SPACES_EVALUATORS_TO_AGGREGATE_LUCoS)
    agg_df_lucos = add_improvability_columns(agg_df_lucos)
    agg_df_lucos = add_rank_columns(agg_df_lucos)
    agg_df_lucos.to_csv(output_path_raw_lucos, index=False)
    check_missing_values(agg_df_lucos)
    format_and_save_agg_excel(agg_df_lucos, output_path_formated_lucos)
    create_results_tables_by_evaluator(agg_df_lucos, AGG_RESULTS_FOLDER, file_prefix="LUCoS")

    logger.info("Saved LUCoS aggregated results to %s", _format_results_path(output_path_raw_lucos))
    logger.info("Saved LUCoS formatted aggregated results to %s", _format_results_path(output_path_formated_lucos))


if __name__ == "__main__":
    main()
